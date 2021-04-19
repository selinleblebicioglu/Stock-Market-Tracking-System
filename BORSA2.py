import requests
from bs4 import BeautifulSoup
import time
from datetime import datetime
import openpyxl
import MAIL

book= openpyxl.load_workbook('C://Users/HP/Desktop/Analiz.xlsx', data_only=True) #verileri işleme soktuğumuz analiz exceli
sheet= book["ASELS"]
ilkdeger= sheet['H2'] #program çalışmadan önceki sinyal durumumuzu kaydediyoruz

book.close()

while True: #sonsuz döngümüze girdik
        try:
                now= datetime.now()
                current_date= now.strftime("%d-%m-%Y") #anlık gün ve saat bilgilerimizi alıyoruz
                current_time= now.strftime("%H:%M:%S")

                if(current_time>="20:15:00"): #program borsanın açık olduğu saatler aralığında çalışıyor
                        break

                elif(current_time>="10:15:00"):

                        book1= openpyxl.load_workbook('C://Users/HP/Desktop/Veri.xlsx')  #bu excel dosyasına verilerimizi kaydediyoruz

                        r= requests.get("http://bigpara.hurriyet.com.tr/borsa/canli-borsa/")  #verimizi aldığımız internet sitesi
                        soup= BeautifulSoup(r.content, "lxml")

                        print("son veri yazma saati: "+current_time)

                        asels_son= soup.find("li", attrs={"id": "h_td_fiyat_id_ASELS"}) #örnek olarak kaydedeceğimiz hisse ASELSAN

                        asels= (current_date, current_time, float(asels_son.text.replace(",",".")))

                        sheet1= book1["ASELS"] #yazdıracağımız excel dosyasındaki sayfanın adı
                        sheet1.append(asels) #yazdıracağımız değişkeni append ile excelin en son satırına ekliyoruz

                        book1.save("C://Users/HP/Desktop/Veri.xlsx")
                        book1.close()

                        time.sleep(5)  #saniye beklemesi koyduk. excelde hesaplamalar yapıldığı için birkaç saniye bekliyoruz

                        book2= openpyxl.load_workbook('C://Users/HP/Desktop/Analiz.xlsx', data_only=True) #analiz excelimizi işlemlerden
                        #sonra tekrar açıyoruz
                        sheet2= book2["ASELS"]
                        sondeger= sheet2['H2'] #işlemlerden sonraki sinyal durumumuzu kaydediyoruz

                        book2.close()

                        if(ilkdeger.value==sondeger.value): #burada önceki ve sonraki sinyal durumumuzu karşılaştırıyoruz
                                #eğer iki sinyal de aynıysa önceden mail gönderme yapmışız demektir. bu yüzden bir işlem gerçekleştirmiyoruz
                                pass

                        elif(ilkdeger.value!=sondeger.value and sondeger.value==1): #fakat önceki ve son sinyallerimiz farklıysa sinyalimiz daha yeni değiştiği için mail
                                #göndermemişiz demektir. bu yüzden bu elif bloğunda mailimizi ilk kez gönderiyoruz
                                sat= 'ASELS SAT',sondeger.value #eğer son sinyal değerimiz 1 ise sat mesajı gönderiyoruz
                                MAIL.mesaj(sat) #mail sayfasındaki mesaj fonksiyonunu çağırdık
                                ilkdeger.value= sondeger.value #ardından programın başında okuduğumuz sinyal durumunu güncelliyoruz.

                        elif(ilkdeger.value!=sondeger.value and sondeger.value==0):
                                al= 'ASELS AL', sondeger.value #eğer son sinyal değerimiz 0 ise al mesajı gönderiyoruz
                                MAIL.mesaj(al)
                                ilkdeger.value= sondeger.value

                        time.sleep(5)  #burada hissenin değerini tekrar okumadan önce bekliyoruz. normalde bekleme süresi daha fazla fakat
                    #örnek olsun diye birkaç saniyelik bekleme koyduk

        except ConnectionResetError: #burada internet sitesine bağlanırken çıkabilecek bir hatayı except bloğu içine yazdık. böylece
                #hatayla karşılaşsa bile program çalışmaya devam edebilir
                pass


